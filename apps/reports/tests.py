

"""Integration tests for the upload/ingest pipeline."""
import io

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse

from apps.accounts.models import Profile
from apps.core.tests_parser import WB_HEADERS

from . import services
from .models import BaseRow, Cost, ExtExpense, ListReport, SelfPurchase

User = get_user_model()


def make_detail_xlsx(rows):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Отчёт (мусорная строка)'])
    ws.append(WB_HEADERS)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'Еженедельный_отчет_692179585.xlsx'
    return buf


def make_list_xlsx():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['№ отчёта', 'Юридическое лицо', 'Дата начала', 'Дата конца',
               'Продажа', 'Стоимость логистики', 'Общая сумма штрафов'])
    ws.append(['692179585', 'ИП Тест', '2025-03-03', '2025-03-09', '1000000', '12345', '100'])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = 'spisok.xlsx'
    return buf


# One Продажа row matching the README reference values.
DETAIL_ROW = [
    'SR1', 'P1', '200000', 'Piccino', 'ART-1', '12345', 'Товар', 'M',
    'Продажа', 'Продажа', '2025-03-07', 1, 772.21, 668.61, 531.71,
    0, 0, '', 'Коледино', 0,
]


class UploadFlowTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('seller', password='pw12345')
        self.profile = Profile.objects.create(user=self.user, name='ИП Тест', tax_type=1, tax_rate=2)
        Cost.objects.create(profile=self.profile, code='12345', cost=235, name='Товар')
        self.client.force_login(self.user)

    def test_detail_upload_computes_and_stores(self):
        f = make_detail_xlsx([DETAIL_ROW])
        resp = self.client.post(reverse('reports:upload_detail'), {'file': f})
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data['ok'])
        self.assertEqual(data['added'], 1)
        self.assertEqual(data['report_num'], '692179585')

        self.assertEqual(BaseRow.objects.filter(profile=self.profile).count(), 1)
        row = BaseRow.objects.get(profile=self.profile)
        self.assertAlmostEqual(row.revenue, 772.21, places=2)
        self.assertAlmostEqual(row.to_transfer, 531.71, places=2)
        self.assertAlmostEqual(row.wb_realized, 668.61, places=2)
        self.assertAlmostEqual(row.cost, 235, places=2)
        self.assertAlmostEqual(row.tax, 13.37, places=2)
        self.assertEqual(row.article, 'ART-1')
        self.assertEqual(row.warehouse, 'Коледино')
        self.assertEqual(row.month, 'Март')

    def test_dedup_by_srid(self):
        self.client.post(reverse('reports:upload_detail'), {'file': make_detail_xlsx([DETAIL_ROW])})
        resp = self.client.post(reverse('reports:upload_detail'), {'file': make_detail_xlsx([DETAIL_ROW])})
        data = resp.json()
        self.assertEqual(data['added'], 0)
        self.assertEqual(data['skipped'], 1)
        self.assertEqual(BaseRow.objects.filter(profile=self.profile).count(), 1)

    def test_same_srid_multiple_rows_all_kept(self):
        # In a WB Подробный отчёт one srid spans several lines (Продажа + Логистика).
        # All must be stored on a single upload — no in-batch srid dedup.
        sale = list(DETAIL_ROW)                      # Продажа, srid SR1
        logistics = list(DETAIL_ROW)
        logistics[8] = ''                            # Тип документа empty
        logistics[9] = 'Логистика'                   # Обоснование = Логистика
        f = make_detail_xlsx([sale, logistics])      # both share srid SR1
        resp = self.client.post(reverse('reports:upload_detail'), {'file': f})
        data = resp.json()
        self.assertEqual(data['added'], 2)           # BOTH kept
        self.assertEqual(data['skipped'], 0)
        self.assertEqual(BaseRow.objects.filter(profile=self.profile).count(), 2)

    def test_list_upload(self):
        resp = self.client.post(reverse('reports:upload_list'), {'file': make_list_xlsx()})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()['ok'])
        lr = ListReport.objects.get(profile=self.profile, report_num='692179585')
        self.assertEqual(lr.legal_entity, 'ИП Тест')
        self.assertAlmostEqual(lr.sale_total, 1000000, places=2)

    def test_recompute_after_cost_change(self):
        self.client.post(reverse('reports:upload_detail'), {'file': make_detail_xlsx([DETAIL_ROW])})
        Cost.objects.filter(profile=self.profile, code='12345').update(cost=300)
        services.recompute_profile(self.profile)
        row = BaseRow.objects.get(profile=self.profile)
        self.assertAlmostEqual(row.cost, 300, places=2)

    def test_bad_extension_rejected(self):
        bad = io.BytesIO(b'hello')
        bad.name = 'note.txt'
        resp = self.client.post(reverse('reports:upload_detail'), {'file': bad})
        self.assertEqual(resp.status_code, 400)
        self.assertFalse(resp.json()['ok'])

    def test_tenant_isolation(self):
        # Another user must not see this profile's rows via the upload page count.
        self.client.post(reverse('reports:upload_detail'), {'file': make_detail_xlsx([DETAIL_ROW])})
        other = User.objects.create_user('other', password='pw12345')
        Profile.objects.create(user=other, name='Other')
        self.client.force_login(other)
        resp = self.client.get(reverse('reports:upload'))
        self.assertEqual(resp.context['base_count'], 0)

class Phase6ToolsTest(TestCase):
    """CRUD wiring for the Себестоимость / Самовыкупы / Внешние расходы pages."""

    def setUp(self):
        self.user = User.objects.create_user('seller', password='pw12345')
        self.profile = Profile.objects.create(user=self.user, name='ИП Тест', tax_type=1, tax_rate=2)
        self.client.force_login(self.user)

    # ---- Себестоимость ----
    def test_cost_add_update_delete_recomputes(self):
        # Ingest one row first so recompute has something to act on.
        Cost.objects.create(profile=self.profile, code='12345', cost=235, name='Товар')
        self.client.post(reverse('reports:upload_detail'), {'file': make_detail_xlsx([DETAIL_ROW])})
        row = BaseRow.objects.get(profile=self.profile)
        self.assertAlmostEqual(row.cost, 235, places=2)

        # Add a new cost via the endpoint.
        resp = self.client.post(reverse('reports:cost_add'),
                                {'article': 'A2', 'code': '999', 'cost': '50'})
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json()['ok'])
        self.assertEqual(Cost.objects.filter(profile=self.profile).count(), 2)

        # Update the matching cost -> base row cost must follow after recompute.
        c = Cost.objects.get(profile=self.profile, code='12345')
        resp = self.client.post(reverse('reports:cost_update'),
                                {'id': c.id, 'field': 'cost', 'value': '300'})
        self.assertTrue(resp.json()['ok'])
        row.refresh_from_db()
        self.assertAlmostEqual(row.cost, 300, places=2)

        # Missing fields are rejected.
        bad = self.client.post(reverse('reports:cost_add'), {'article': '', 'code': ''})
        self.assertEqual(bad.status_code, 400)

        # Delete -> back to no cost for that code.
        self.client.post(reverse('reports:cost_delete'), {'id': c.id})
        row.refresh_from_db()
        self.assertAlmostEqual(row.cost, 0, places=2)

    def test_costs_page_lists_missing_codes(self):
        Cost.objects.create(profile=self.profile, code='12345', cost=235, name='Товар')
        self.client.post(reverse('reports:upload_detail'), {'file': make_detail_xlsx([DETAIL_ROW])})
        resp = self.client.get(reverse('reports:costs'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['total'], 1)

    # ---- Самовыкупы ----
    def test_self_purchase_save_and_delete(self):
        resp = self.client.post(reverse('reports:self_purchase_save'),
                                {'srid': 'S1', 'cost': '100', 'logistics': '20',
                                 'date': '2025-03-07', 'note': 'n'})
        self.assertTrue(resp.json()['ok'])
        sp = SelfPurchase.objects.get(profile=self.profile)
        self.assertAlmostEqual(sp.total, 120, places=2)

        # Edit it.
        self.client.post(reverse('reports:self_purchase_save'),
                         {'id': sp.id, 'srid': 'S1', 'cost': '200', 'logistics': '0'})
        sp.refresh_from_db()
        self.assertAlmostEqual(sp.total, 200, places=2)

        # SRID required.
        self.assertEqual(self.client.post(reverse('reports:self_purchase_save'),
                                          {'srid': ''}).status_code, 400)

        self.client.post(reverse('reports:self_purchase_delete'), {'id': sp.id})
        self.assertFalse(SelfPurchase.objects.filter(id=sp.id).exists())

    def test_self_purchase_feeds_formula(self):
        # A самовыкуп for the same srid as a base row should populate AL (samovikup_cost).
        Cost.objects.create(profile=self.profile, code='12345', cost=235, name='Товар')
        self.client.post(reverse('reports:upload_detail'), {'file': make_detail_xlsx([DETAIL_ROW])})
        row = BaseRow.objects.get(profile=self.profile)
        self.client.post(reverse('reports:self_purchase_save'),
                         {'srid': row.srid, 'cost': '111', 'logistics': '0'})
        row.refresh_from_db()
        self.assertAlmostEqual(row.samovikup_cost, 111, places=2)

    # ---- Внешние расходы ----
    def test_ext_expense_crud(self):
        resp = self.client.post(reverse('reports:ext_expense_save'),
                                {'description': 'Реклама IG', 'category': 'Реклама',
                                 'amount': '5000', 'date': '2026-05-15'})
        self.assertTrue(resp.json()['ok'])
        e = ExtExpense.objects.get(profile=self.profile)
        self.assertEqual(e.month, '2026-05')

        # Tavsif + summa required.
        self.assertEqual(self.client.post(reverse('reports:ext_expense_save'),
                                          {'description': '', 'amount': ''}).status_code, 400)

        # Month filter on the page.
        resp = self.client.get(reverse('reports:ext_expenses'), {'month': '2026-05'})
        self.assertEqual(len(resp.context['items']), 1)
        resp = self.client.get(reverse('reports:ext_expenses'), {'month': '2099-01'})
        self.assertEqual(len(resp.context['items']), 0)

        self.client.post(reverse('reports:ext_expense_delete'), {'id': e.id})
        self.assertFalse(ExtExpense.objects.filter(id=e.id).exists())

    # ---- Settings ----
    def test_settings_save_recomputes_on_tax_change(self):
        Cost.objects.create(profile=self.profile, code='12345', cost=235, name='Товар')
        self.client.post(reverse('reports:upload_detail'), {'file': make_detail_xlsx([DETAIL_ROW])})
        resp = self.client.post(reverse('accounts:settings_save'), {
            'name': 'ИП Тест', 'currency': '₽', 'tax_type': '3', 'tax_rate': '0'})
        self.assertEqual(resp.status_code, 302)
        self.profile.refresh_from_db()
        self.assertEqual(self.profile.tax_type, 3)
        # tax_type 3 = Не считать -> tax becomes 0 on every row.
        self.assertEqual(BaseRow.objects.filter(profile=self.profile).exclude(tax=0).count(), 0)

    def test_tenant_isolation_on_costs(self):
        Cost.objects.create(profile=self.profile, code='1', cost=1, name='mine')
        other = User.objects.create_user('other2', password='pw12345')
        Profile.objects.create(user=other, name='Other')
        self.client.force_login(other)
        resp = self.client.get(reverse('reports:costs'))
        self.assertEqual(resp.context['total'], 0)


def make_cap_stock_xlsx():
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(['Бренд', 'Предмет', 'Артикул продавца', 'Артикул WB', 'Баркод',
               'Размер вещи', 'В пути до получателей', 'В пути возвраты',
               'Всего находится на складах', 'Коледино', 'Электросталь'])
    ws.append(['Piccino', 'Костюм', 'ART-1', '12345', 'BC1', 'M', 0, 2, 10, 6, 4])
    ws.append(['Piccino', 'Халат', 'ART-2', '67890', 'BC2', 'L', 0, 0, 5, 5, 0])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); buf.name = 'ostatki.xlsx'
    return buf


def make_cap_sales_xlsx():
    import openpyxl
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(['Отчёт по продажам (title)'])  # WB sales files have a title row first
    ws.append(['Бренд', 'Предмет', 'Артикул продавца', 'Артикул WB', 'Баркод',
               'Размер', 'Склад', 'шт.', 'Сумма заказов, руб.', 'Выкупили, шт.',
               'К перечислению', 'Текущий остаток'])
    ws.append(['Piccino', 'Костюм', 'ART-1', '12345', 'BC1', 'M', 'Коледино', 14, 14000, 13, 9000, 6])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0); buf.name = 'prodaji.xlsx'
    return buf


class CapitalizationTest(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('capuser', password='pw12345')
        self.profile = Profile.objects.create(user=self.user, name='ИП', tax_type=1, tax_rate=2)
        self.client.force_login(self.user)

    def test_stock_upload_and_parse(self):
        resp = self.client.post(reverse('reports:cap_upload'),
                                {'kind': 'stock', 'file': make_cap_stock_xlsx()})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()['count'], 2)
        from .models import CapStock
        rows = list(CapStock.objects.filter(profile=self.profile))
        self.assertEqual(len(rows), 2)
        bc1 = next(r for r in rows if r.barcode == 'BC1')
        self.assertEqual(bc1.raw['warehouses'], {'Коледино': 6, 'Электросталь': 4})
        self.assertEqual(bc1.raw['total_in_warehouses'], 10)
        self.assertEqual(bc1.raw['code'], '12345')

    def test_sales_upload(self):
        resp = self.client.post(reverse('reports:cap_upload'),
                                {'kind': 'sales', 'file': make_cap_sales_xlsx()})
        self.assertEqual(resp.json()['count'], 1)
        from .models import CapSales
        s = CapSales.objects.get(profile=self.profile)
        self.assertEqual(s.barcode, 'BC1')
        self.assertEqual(s.raw['ordered'], 14)
        self.assertEqual(s.raw['sklad'], 'Коледино')

    def test_upload_reupload_replaces(self):
        self.client.post(reverse('reports:cap_upload'), {'kind': 'stock', 'file': make_cap_stock_xlsx()})
        self.client.post(reverse('reports:cap_upload'), {'kind': 'stock', 'file': make_cap_stock_xlsx()})
        from .models import CapStock
        self.assertEqual(CapStock.objects.filter(profile=self.profile).count(), 2)  # replaced, not doubled

    def test_params_save_and_clear(self):
        resp = self.client.post(reverse('reports:cap_params_save'),
                                {'days_of_sales': '14', 'delivery_days': '10',
                                 'reserve_days': '21', 'delivery_cost': '50',
                                 'selected_warehouse': 'Коледино'})
        self.assertTrue(resp.json()['ok'])
        from .models import CapParams
        p = CapParams.objects.get(profile=self.profile)
        self.assertEqual(p.days_of_sales, 14)
        self.assertEqual(p.reserve_days, 21)
        self.assertEqual(p.selected_warehouse, 'Коледино')

        self.client.post(reverse('reports:cap_upload'), {'kind': 'stock', 'file': make_cap_stock_xlsx()})
        resp = self.client.post(reverse('reports:cap_clear'), {'kind': 'stock'})
        self.assertTrue(resp.json()['ok'])
        from .models import CapStock
        self.assertEqual(CapStock.objects.filter(profile=self.profile).count(), 0)

    def test_cost_set_by_barcode_recomputes(self):
        # Base row tied to wb_article 12345 → its AC follows the cost edit.
        Cost.objects.create(profile=self.profile, code='12345', cost=100, name='Костюм')
        services.ingest_detail(self.profile, {'sales': [{
            'srid': 'X', 'operation_type': 'Продажа', 'operation_reason': 'Продажа', 'qty': 1,
            'price_with_discount': 772.21, 'revenue_wb': 668.61, 'to_pay': 531.71,
            'wb_article': '12345', 'sale_date': '2025-03-07'}],
            'info': {'filename': 'r.xlsx', 'reportNum': '692179585'}})
        # Set a barcode-keyed cost (no barcode match → clones the code sample)
        resp = self.client.post(reverse('reports:cap_cost_set'),
                                {'barcode': 'BC1', 'code': '12345', 'cost': '250'})
        self.assertTrue(resp.json()['ok'])
        # A barcode-keyed Cost now exists; base recompute still uses code 12345 cost (100).
        self.assertTrue(Cost.objects.filter(profile=self.profile, barcode='BC1').exists())

    def test_page_loads_and_tenant_isolated(self):
        self.client.post(reverse('reports:cap_upload'), {'kind': 'stock', 'file': make_cap_stock_xlsx()})
        resp = self.client.get(reverse('analytics:capitalization'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context['cap_stock_json']), 2)
        other = User.objects.create_user('capother', password='pw12345')
        Profile.objects.create(user=other, name='Other')
        self.client.force_login(other)
        resp = self.client.get(reverse('analytics:capitalization'))
        self.assertEqual(len(resp.context['cap_stock_json']), 0)
