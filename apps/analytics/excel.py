"""Shared .xlsx report builder used by the monthly and per-report (weekly)
download buttons. Given a KPI dict + per-product rows (the outputs of
analytics.services.compute_kpi / by_product) it builds a two-sheet workbook:

  * "Hisobot"  — the KPI / expense breakdown summary
  * "Tovarlar" — per-product sales & profit/loss table with a JAMI row
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_MONEY_FMT = '#,##0'
_PCT_FMT = '0.0"%"'


def build_report_workbook(title, kpi, products, currency, subtitle=''):
    """Return an openpyxl Workbook for one period (month or weekly report)."""
    navy = PatternFill('solid', fgColor='042C53')
    blue = PatternFill('solid', fgColor='185FA5')
    gray = PatternFill('solid', fgColor='F1F3F5')
    white_bold = Font(bold=True, color='FFFFFF')
    bold = Font(bold=True)
    thin = Side(style='thin', color='D0D7DE')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal='right')

    wb = Workbook()

    # ===== Sheet 1: Hisobot (KPI summary / xarajatlar) =====
    ws = wb.active
    ws.title = 'Hisobot'
    ws.merge_cells('A1:B1')
    head = ws['A1']
    head.value = title
    head.font = white_bold
    head.fill = navy
    head.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 26
    ws['B2'] = subtitle or currency
    ws['B2'].alignment = right
    ws['B2'].font = Font(italic=True, color='888888')

    summary = [
        ('Sotilgan tovarlar (dona)', kpi['sales_qty'], None),
        ('Daromad', kpi['revenue'], _MONEY_FMT),
        ('Oplata RS (Kʻ perech.)', kpi['to_pay_rs'], _MONEY_FMT),
        ('WB ushlanmalari', -kpi['all_wb_deductions'], _MONEY_FMT),
        ('  · Komissiya', -kpi['commission'], _MONEY_FMT),
        ('  · Ekvayring', -kpi['acquiring'], _MONEY_FMT),
        ('  · Logistika', -kpi['logistics_total'], _MONEY_FMT),
        ('  · Saqlash', -kpi['storage'], _MONEY_FMT),
        ('  · Jarima', -kpi['fines'], _MONEY_FMT),
        ('  · Reklama', -kpi['promotion'], _MONEY_FMT),
        ('Tannarx', -kpi['cost'], _MONEY_FMT),
        ('Soliq', -kpi['tax'], _MONEY_FMT),
        ('Tashqi xarajatlar', -kpi['ext_expenses'], _MONEY_FMT),
        ('Toza foyda', kpi['net_profit'], _MONEY_FMT),
        ('Marja', kpi['margin_pct'], _PCT_FMT),
        ('ROI', kpi['roi_pct'], _PCT_FMT),
    ]
    row = 3
    for label, value, fmt in summary:
        c_label = ws.cell(row=row, column=1, value=label)
        c_value = ws.cell(row=row, column=2, value=round(value, 2) if value else 0)
        c_value.alignment = right
        if fmt:
            c_value.number_format = fmt
        if label in ('Daromad', 'WB ushlanmalari', 'Tannarx', 'Toza foyda'):
            c_label.font = bold
            c_value.font = bold
        if label == 'Toza foyda':
            fill = PatternFill('solid', fgColor='D7F5DD' if kpi['net_profit'] >= 0 else 'FBDDDD')
            c_label.fill = fill
            c_value.fill = fill
        row += 1
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18

    # ===== Sheet 2: Tovarlar (per-product profit/loss) =====
    ws2 = wb.create_sheet('Tovarlar')
    headers = ['Artikul', 'Nomi', 'Brend', 'Sotilgan', 'Qaytarilgan', 'Daromad',
               'Komissiya', 'Logistika', 'Saqlash', 'Reklama', 'Tannarx', 'Soliq',
               'Foyda/Zarar', 'Marja %', 'ROI %']
    for col, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = white_bold
        c.fill = blue
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws2.row_dimensions[1].height = 28

    for i, p in enumerate(products, start=2):
        vals = [
            p.get('article') or '—', p.get('product') or '', p.get('brand') or '',
            p.get('buyouts_qty') or 0, p.get('returns_qty') or 0,
            round(p.get('revenue') or 0), round(p.get('commission') or 0),
            round(p.get('logistics') or 0), round(p.get('storage') or 0),
            round(p.get('promotion') or 0), round(p.get('cost') or 0),
            round(p.get('tax') or 0), round(p.get('profit') or 0),
            round(p.get('margin_pct') or 0, 1), round(p.get('roi_pct') or 0, 1),
        ]
        for col, v in enumerate(vals, start=1):
            c = ws2.cell(row=i, column=col, value=v)
            c.border = border
            if col >= 4:
                c.alignment = right
            if col in range(6, 13):
                c.number_format = _MONEY_FMT
            elif col == 13:  # Foyda/Zarar
                c.number_format = _MONEY_FMT
                c.font = Font(bold=True, color='1A7F37' if (p.get('profit') or 0) >= 0 else 'CF222E')
            elif col in (14, 15):
                c.number_format = _PCT_FMT
        if i % 2 == 0:
            for col in range(1, len(headers) + 1):
                if ws2.cell(row=i, column=col).fill.patternType is None:
                    ws2.cell(row=i, column=col).fill = gray

    # totals row
    if products:
        tr = len(products) + 2
        ws2.cell(row=tr, column=1, value='JAMI').font = bold
        totals = {
            4: sum(p.get('buyouts_qty') or 0 for p in products),
            5: sum(p.get('returns_qty') or 0 for p in products),
            6: round(sum(p.get('revenue') or 0 for p in products)),
            7: round(sum(p.get('commission') or 0 for p in products)),
            8: round(sum(p.get('logistics') or 0 for p in products)),
            9: round(sum(p.get('storage') or 0 for p in products)),
            10: round(sum(p.get('promotion') or 0 for p in products)),
            11: round(sum(p.get('cost') or 0 for p in products)),
            12: round(sum(p.get('tax') or 0 for p in products)),
            13: round(sum(p.get('profit') or 0 for p in products)),
        }
        for col, v in totals.items():
            c = ws2.cell(row=tr, column=col, value=v)
            c.font = bold
            c.alignment = right
            c.fill = gray
            if col >= 6:
                c.number_format = _MONEY_FMT
        for col in (1, 2, 3, 14, 15):
            ws2.cell(row=tr, column=col).fill = gray

    widths = [16, 30, 14, 10, 12, 14, 12, 12, 11, 11, 13, 11, 14, 9, 9]
    for col, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = 'A2'

    return wb